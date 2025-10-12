from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages

from django.contrib.auth.models import User

# --- BU SATIRI GÜNCELLEYİN ---
from datetime import datetime, date, timedelta
import calendar
from collections import defaultdict

from .models import Doktor, IzinTalebi, Nobet, HastaneAyarlari
from .nobet_planner import NobetPlanlayici

import pandas as pd
import io
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, PatternFill

# --- YÖNLENDİRME FONKSİYONLARI ---

@login_required
def doktor_paneli_redirect(request):
    bugun = date.today()
    return redirect('doktor_paneli', yil=bugun.year, ay=bugun.month)

@login_required
def takvim_redirect(request):
    bugun = date.today()
    return redirect('takvim_gorunumu', yil=bugun.year, ay=bugun.month)

# --- ANA FONKSİYONLAR (GÜNCELLENDİ) ---

@login_required
def doktor_paneli(request, yil, ay):
    if request.user.is_staff:
        return redirect('yonetim_paneli')

    try:
        doktor = request.user.doktor
    except Doktor.DoesNotExist:
        messages.error(request, "Doktor profiliniz bulunamadı.")
        return redirect('login')

    if request.method == 'POST':
        tarih_str = request.POST.get('tarih')
        if tarih_str:
            izin_tarihi = date.fromisoformat(tarih_str)
            ayarlar = HastaneAyarlari.get_solo()
            IZIN_LIMITI = ayarlar.aylik_izin_limiti
            mevcut_izin_sayisi = IzinTalebi.objects.filter(doktor=doktor, tarih__year=izin_tarihi.year, tarih__month=izin_tarihi.month).count()
            if mevcut_izin_sayisi >= IZIN_LIMITI:
                messages.error(request, f"{izin_tarihi.strftime('%B')} ayı için izin limitinizi ({IZIN_LIMITI} gün) doldurdunuz.")
            else:
                nesne, olusturuldu = IzinTalebi.objects.get_or_create(doktor=doktor, tarih=izin_tarihi)
                if olusturuldu:
                    messages.success(request, f"{izin_tarihi.strftime('%d %B %Y')} tarihli izin talebiniz eklendi.")
                else:
                    messages.warning(request, "Bu tarih için zaten bir talebiniz mevcut.")
        return redirect('doktor_paneli', yil=yil, ay=ay)

    mevcut_tarih = date(yil, ay, 1)
    onceki_ay_tarih = mevcut_tarih - timedelta(days=1)
    sonraki_ay_tarih = (mevcut_tarih + timedelta(days=32)).replace(day=1) # Bir sonraki ayın ilk gününü bulmak için daha güvenli yöntem
    
    ay_isimleri = ["", "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran", "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]
    
    mevcut_izinler = IzinTalebi.objects.filter(doktor=doktor, tarih__year=yil, tarih__month=ay).order_by('tarih')

    context = {
        'doktor': doktor, 'mevcut_izinler': mevcut_izinler,
        'yil': yil, 'ay': ay,
        'gosterilecek_ay_ismi': ay_isimleri[ay],
        'onceki_yil': onceki_ay_tarih.year, 'onceki_ay': onceki_ay_tarih.month,
        'sonraki_yil': sonraki_ay_tarih.year, 'sonraki_ay': sonraki_ay_tarih.month,
    }
    return render(request, 'core/doktor_paneli.html', context)

@login_required
def izin_sil(request, izin_id):
    izin = get_object_or_404(IzinTalebi, id=izin_id)
    silinen_yil = izin.tarih.year
    silinen_ay = izin.tarih.month
    if izin.doktor.user == request.user:
        if request.method == 'POST':
            izin.delete()
            messages.success(request, f"{izin.tarih.strftime('%d %B %Y')} tarihli izin talebiniz başarıyla silindi.")
    else:
        messages.error(request, "Bu işlemi yapma yetkiniz yok.")
    return redirect('doktor_paneli', yil=silinen_yil, ay=silinen_ay)

@login_required
def takvim_gorunumu(request, yil, ay):
    nobetler_qs = Nobet.objects.filter(tarih__year=yil, tarih__month=ay).select_related('doktor')
    gunlere_gore_nobetler = defaultdict(list)
    for nobet in nobetler_qs:
        gunlere_gore_nobetler[nobet.tarih.day].append(nobet)

    takvim_cal = calendar.Calendar()
    ay_gunleri = takvim_cal.itermonthdays2(yil, ay)
    ay_isimleri = ["", "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran", "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]
    
    mevcut_tarih = date(yil, ay, 1)
    onceki_ay_tarih = mevcut_tarih - timedelta(days=1)
    sonraki_ay_tarih = (mevcut_tarih + timedelta(days=32)).replace(day=1) # Daha güvenli yöntem

    context = {
        'yil': yil, 'ay': ay, 'ay_ismi': ay_isimleri[ay],
        'ay_gunleri': ay_gunleri, 'nobetler': gunlere_gore_nobetler,
        'onceki_yil': onceki_ay_tarih.year, 'onceki_ay': onceki_ay_tarih.month,
        'sonraki_yil': sonraki_ay_tarih.year, 'sonraki_ay': sonraki_ay_tarih.month,
    }
    return render(request, 'core/takvim.html', context)

@staff_member_required
def yonetim_paneli(request):
    if request.method == 'POST':
        try:
            yil = int(request.POST.get('yil'))
            ay = int(request.POST.get('ay'))
            planlayici = NobetPlanlayici(yil, ay)
            plan = planlayici.plani_olustur()
            
            if not plan:
                messages.error(request, f"{yil}-{ay} için yeterli doktor veya geçerli bir kombinasyon bulunamadığından plan oluşturulamadı.")
            else:
                planlayici.plani_kaydet(plan)
                messages.success(request, f"{yil} yılının {ay}. ayı için nöbet planı başarıyla oluşturuldu!")
            
            return redirect('yonetim_paneli')
        except Exception as e:
            messages.error(request, f"Bir hata oluştu: {e}")
            return redirect('yonetim_paneli')

    simdiki_zaman = datetime.now()
    gelecek_ay = (simdiki_zaman.month % 12) + 1
    gelecek_yil = simdiki_zaman.year + (1 if simdiki_zaman.month == 12 else 0)
    context = {
        'varsayilan_yil': gelecek_yil, 'varsayilan_ay': gelecek_ay,
    }
    return render(request, 'core/yonetim_paneli.html', context)

@login_required
def export_takvim_excel(request, yil, ay):
    nobetler_qs = Nobet.objects.filter(tarih__year=yil, tarih__month=ay).select_related('doktor').order_by('tarih')

    if not nobetler_qs.exists():
        messages.warning(request, f"{yil}-{ay} dönemi için dışa aktarılacak nöbet bulunamadı.")
        return redirect('takvim_gorunumu', yil=yil, ay=ay)

    data = []
    for nobet in nobetler_qs:
        data.append({
            'Tarih': nobet.tarih, 'Doktor Adı': nobet.doktor.ad_soyad,
            'Kıdem': nobet.doktor.get_kidem_display(), 'Bölüm': nobet.get_bolum_display().replace(' Alan', ''),
            'Zorunlu': nobet.izin_iptal_edildi
        })
    df_raw = pd.DataFrame(data)
    df_raw['Tarih'] = pd.to_datetime(df_raw['Tarih'])

    df_raw['Doktor Adı'] = df_raw.apply(lambda row: f"{row['Doktor Adı']} (Zorunlu!)" if row['Zorunlu'] else row['Doktor Adı'], axis=1)
    df_raw['Pozisyon'] = df_raw.groupby(['Tarih', 'Bölüm']).cumcount() + 1
    df_raw['Pozisyon'] = df_raw['Bölüm'] + ' ' + df_raw['Pozisyon'].astype(str)
    df_raw['Pozisyon'] = df_raw['Pozisyon'].str.replace(r' 1$', '', regex=True)
    
    df_pivot = df_raw.pivot_table(index='Tarih', columns='Pozisyon', values='Doktor Adı', aggfunc='first').fillna('')
    sutun_sirasi = ['Kırmızı', 'Sarı', 'Yeşil', 'Yeşil 2', 'Yeşil 3']
    mevcut_sutunlar = [s for s in sutun_sirasi if s in df_pivot.columns]
    df_pivot = df_pivot[mevcut_sutunlar]

    gun_isimleri_tr = ["Pzt", "Salı", "Çar", "Per", "Cuma", "Cmt", "Paz"]
    df_pivot.index = df_pivot.index.strftime('%d.%m.%Y') + ' - ' + df_pivot.index.weekday.map(lambda x: gun_isimleri_tr[x])

    df_raw['Hafta Sonu'] = df_raw['Tarih'].dt.weekday >= 5
    df_stats = df_raw.groupby('Doktor Adı').agg(
        Toplam_Nöbet=('Doktor Adı', 'size'), Hafta_Sonu_Nöbeti=('Hafta Sonu', 'sum'),
        Zorunlu_Atama_Sayısı=('Zorunlu', 'sum')
    ).astype(int)
    bolum_stats = df_raw.groupby(['Doktor Adı', 'Bölüm']).size().unstack(fill_value=0)
    df_stats = pd.concat([df_stats, bolum_stats], axis=1).fillna(0).astype(int)
    df_stats = df_stats.sort_values(by='Toplam_Nöbet', ascending=False)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_pivot.to_excel(writer, sheet_name='Nöbet Takvimi')
        df_stats.to_excel(writer, sheet_name='Özet & İstatistikler')

        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="004C99", end_color="004C99", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center')

        ws_takvim = writer.sheets['Nöbet Takvimi']
        ws_takvim.column_dimensions['A'].width = 22
        for col_letter in ['B', 'C', 'D', 'E', 'F']:
             ws_takvim.column_dimensions[col_letter].width = 25
        for cell in ws_takvim[1]:
            cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align
        for row in ws_takvim.iter_rows(min_row=2):
            row[0].alignment = left_align

        ws_stats = writer.sheets['Özet & İstatistikler']
        ws_stats.column_dimensions['A'].width = 25
        for cell in ws_stats[1]:
            cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align

    output.seek(0)
    filename = f"Nobet_Takvimi_{yil}_{ay}.xlsx"
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

# --- GEÇİCİ YÖNETİCİ OLUŞTURMA FONKSİYONU ---
def create_superuser_gecici(request):
    ADMIN_USERNAME = 'admin'  # Canlı site için admin kullanıcı adın
    ADMIN_PASSWORD = 'mMavisuluk23' # MUTLAKA GÜÇLÜ VE UNUTMAYACAĞIN BİR ŞİFRE SEÇ!
    ADMIN_EMAIL = 'berkoner05@gmail.com' # Kendi e-posta adresini yazabilirsin

    # Kullanıcı zaten var mı diye kontrol et, varsa tekrar oluşturma
    if not User.objects.filter(username=ADMIN_USERNAME).exists():
        User.objects.create_superuser(ADMIN_USERNAME, ADMIN_EMAIL, ADMIN_PASSWORD)
        return HttpResponse("<h1>Yönetici hesabı başarıyla oluşturuldu!</h1><p>Artık bu sayfayı kapatabilir ve bu kodu projenizden silebilirsiniz.</p>")
    else:
        return HttpResponse("<h1>Yönetici hesabı zaten mevcut.</h1>")